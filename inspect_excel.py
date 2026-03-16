"""Inspect Excel template structure for replication."""
import sys
try:
    import openpyxl
except ImportError:
    print("openpyxl not installed")
    sys.exit(1)

path = r"c:\Users\Work\Desktop\Metro Repo\Metro GL3 (November - 2025).xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    print("=== SHEET:", repr(name), "===")
    print("Max row:", ws.max_row, "Max col:", ws.max_column)
    merges = list(ws.merged_cells.ranges)
    if merges:
        print("Merged:", [str(m) for m in merges])
    for r in range(1, min(50, ws.max_row + 1)):
        row_vals = []
        for c in range(1, min(16, ws.max_column + 1)):
            v = ws.cell(r, c).value
            row_vals.append(repr(v)[:55] if v is not None else "")
        print(r, "|", " | ".join(row_vals))
    print()

wb.close()
